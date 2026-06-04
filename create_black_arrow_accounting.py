from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define colors
PRIMARY_COLOR = "0F3460"  # Navy
ACCENT_COLOR = "F59E0B"   # Gold
HEADER_FILL = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUMMARY_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
SUMMARY_FONT = Font(bold=True, size=10)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== SHEET 1: PROJECT COSTING ====================
ws1 = wb.create_sheet("Project Costing", 0)

# Set column widths
col_widths = [5, 12, 15, 15, 15, 5, 18, 12, 15, 15, 12, 12, 15, 5, 15, 15, 18, 12, 20]
for i, width in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# Headers for Sheet 1
headers_1 = [
    "SN", "INV Ref", "Sales", "Output VAT (15%)", "Total Sales", " ",
    "Material Cost", "Input VAT (15%)", "Transportation", "Transport VAT (15%)",
    "Total COGS", "P.O No", "Supplier ref", " ", "GROSS PROFIT", "VAT Difference",
    "Company Expense VAT", "FINAL VAT", "ZYVC FINAL TOTAL"
]

# Add headers
for col, header in enumerate(headers_1, 1):
    cell = ws1.cell(row=1, column=col)
    cell.value = header
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws1.row_dimensions[1].height = 30

# Add sample data row
sample_data = [1, "INV-001", 50000, None, None, "", 30000, None, 2000, None, None, "PO-001", "Supplier A", "", None, None, None, None, None]

for col, value in enumerate(sample_data, 1):
    cell = ws1.cell(row=2, column=col)
    cell.value = value
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add number formatting for amounts
    if col in [3, 4, 5, 7, 8, 9, 10, 11, 15, 16, 17, 18, 19]:
        cell.number_format = '#,##0.00'

# Add formulas for row 2
ws1.cell(row=2, column=4).value = f"=C2*0.15"  # Output VAT
ws1.cell(row=2, column=5).value = f"=C2+D2"    # Total Sales
ws1.cell(row=2, column=8).value = f"=G2*0.15"  # Input VAT
ws1.cell(row=2, column=10).value = f"=I2*0.15" # Transport VAT
ws1.cell(row=2, column=11).value = f"=G2+H2+I2+J2"  # Total COGS
ws1.cell(row=2, column=15).value = f"=E2-K2"   # GROSS PROFIT (Total Sales - Total COGS)
ws1.cell(row=2, column=16).value = f"=D2-(H2+J2)"   # VAT Difference

# ==================== SHEET 2: COMPANY EXPENSE ====================
ws2 = wb.create_sheet("Company Expense", 1)

# Set column widths
col_widths_2 = [5, 12, 18, 15, 20, 15, 12, 15, 15]
for i, width in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# Headers for Sheet 2
headers_2 = [
    "SN", "Date", "Invoice Reference", "Item", "Description",
    "Paid By", "Expense", "Output VAT (15%)", "Total Expenses"
]

# Add headers
for col, header in enumerate(headers_2, 1):
    cell = ws2.cell(row=1, column=col)
    cell.value = header
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws2.row_dimensions[1].height = 30

# Add 30 data rows
for row in range(2, 32):
    for col in range(1, 10):
        cell = ws2.cell(row=row, column=col)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Add formulas for VAT and Total
        if col == 8:  # Output VAT
            cell.value = f"=G{row}*0.15"
            cell.number_format = '#,##0.00'
        elif col == 9:  # Total Expenses
            cell.value = f"=G{row}+H{row}"
            cell.number_format = '#,##0.00'
        elif col == 7:  # Expense column
            cell.number_format = '#,##0.00'
        elif col == 2:  # Date column
            cell.number_format = 'mm/dd/yyyy'

# Add summary row at the bottom
summary_row = 33
ws2.row_dimensions[summary_row].height = 25

# Summary labels and values
ws2.cell(row=summary_row, column=1).value = "TOTAL"
ws2.cell(row=summary_row, column=1).font = SUMMARY_FONT
ws2.cell(row=summary_row, column=1).fill = SUMMARY_FILL
ws2.cell(row=summary_row, column=1).border = BORDER

# Total Expense formula
total_expense_cell = ws2.cell(row=summary_row, column=7)
total_expense_cell.value = f"=SUM(G2:G31)"
total_expense_cell.font = SUMMARY_FONT
total_expense_cell.fill = SUMMARY_FILL
total_expense_cell.border = BORDER
total_expense_cell.number_format = '#,##0.00'

# Total VAT formula
total_vat_cell = ws2.cell(row=summary_row, column=8)
total_vat_cell.value = f"=SUM(H2:H31)"
total_vat_cell.font = SUMMARY_FONT
total_vat_cell.fill = SUMMARY_FILL
total_vat_cell.border = BORDER
total_vat_cell.number_format = '#,##0.00'

# Total Expenses formula
total_all_cell = ws2.cell(row=summary_row, column=9)
total_all_cell.value = f"=SUM(I2:I31)"
total_all_cell.font = SUMMARY_FONT
total_all_cell.fill = SUMMARY_FILL
total_all_cell.border = BORDER
total_all_cell.number_format = '#,##0.00'

# Save the file
file_path = r"C:\Users\Afzal\Downloads\Black_Arrow_Accounting.xlsx"
wb.save(file_path)
print(f"Excel file created successfully: {file_path}")
