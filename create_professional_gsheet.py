from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Define professional colors
NAVY = "0F3460"
GOLD = "F59E0B"
LIGHT_GRAY = "F3F4F6"
DARK_GRAY = "E5E7EB"
WHITE = "FFFFFF"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(bold=True, color=WHITE, size=11)
SECTION_FILL = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
SECTION_FONT = Font(bold=True, color=WHITE, size=10)
TOTAL_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
TOTAL_FONT = Font(bold=True, size=10, color=NAVY)
BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# ========================== SHEET 1: EXPENSE ==========================
ws1 = wb.create_sheet("Expense", 0)

# Set column widths
widths_1 = [12, 18, 15, 20, 18, 15, 15, 15]
for i, width in enumerate(widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# Headers
headers_1 = ["Date", "Invoice Ref", "Item", "Description", "Paid By", "Expense", "Output VAT (15%)", "Total Expense"]
for col, header in enumerate(headers_1, 1):
    cell = ws1.cell(row=1, column=col)
    cell.value = header
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws1.row_dimensions[1].height = 25

# Add 50 data rows
sar_format = '"SAR"#,##0.00'
date_format = 'mm/dd/yyyy'

for row in range(2, 52):
    for col in range(1, 9):
        cell = ws1.cell(row=row, column=col)
        cell.border = BORDER

        if col == 1:  # Date
            cell.number_format = date_format
        elif col == 6:  # Expense (user enters)
            cell.number_format = sar_format
        elif col == 7:  # Output VAT = Expense * 0.15
            cell.value = f"=F{row}*0.15"
            cell.number_format = sar_format
        elif col == 8:  # Total Expense = Expense + Output VAT
            cell.value = f"=F{row}+G{row}"
            cell.number_format = sar_format
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Add TOTAL row
total_row = 52
ws1.cell(row=total_row, column=1).value = "TOTAL"
ws1.cell(row=total_row, column=1).fill = TOTAL_FILL
ws1.cell(row=total_row, column=1).font = TOTAL_FONT
ws1.cell(row=total_row, column=1).border = BORDER

for col in range(6, 9):
    cell = ws1.cell(row=total_row, column=col)
    col_letter = get_column_letter(col)
    cell.value = f"=SUM({col_letter}2:{col_letter}51)"
    cell.number_format = sar_format
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    cell.border = BORDER

# ========================== SHEET 2: PROJECT COSTING ==========================
ws2 = wb.create_sheet("Project Costing", 1)

# Set column widths - extended for all sections
widths_2 = [5, 15, 15, 15, 15, 3, 15, 15, 15, 15, 15, 15, 15, 3, 15, 15, 15, 15, 15, 15]
for i, width in enumerate(widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# Headers with section breaks
headers_2 = [
    "SN",
    "Invoice Ref",
    "Sales",
    "Output VAT",
    "Total Sales",
    " ",  # Space
    "Material Cost",
    "Input VAT",
    "Transportation",
    "Transport VAT",
    "Total COGS",
    "P.O No",
    "Supplier Ref",
    " ",  # Space
    "Gross Profit",
    "VAT Difference",
    "Zakat @2.5%",
    "Yousef @7.5%",
    "Everyone @17%",
    "Charity @5%"
]

# Add headers with color-coding by section
for col, header in enumerate(headers_2, 1):
    cell = ws2.cell(row=1, column=col)
    cell.value = header
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # SALES section (columns 2-5): Gold
    if col in [2, 3, 4, 5]:
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
    # PURCHASE section (columns 7-13): Navy
    elif col in [7, 8, 9, 10, 11, 12, 13]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    # SUMMARY section (columns 15-20): Gold
    elif col in [15, 16, 17, 18, 19, 20]:
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
    # Spacers: Light gray
    else:
        cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

ws2.row_dimensions[1].height = 25

# Add 50 data rows with formulas
for row in range(2, 52):
    for col in range(1, 21):
        cell = ws2.cell(row=row, column=col)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center")

        if col == 1:  # SN
            cell.value = row - 1
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # SALES SECTION
        elif col == 3:  # Sales (user enters)
            cell.number_format = sar_format
        elif col == 4:  # Output VAT = Sales * 0.15
            cell.value = f"=C{row}*0.15"
            cell.number_format = sar_format
        elif col == 5:  # Total Sales = Sales + Output VAT
            cell.value = f"=C{row}+D{row}"
            cell.number_format = sar_format

        # PURCHASE SECTION
        elif col == 7:  # Material Cost (user enters)
            cell.number_format = sar_format
        elif col == 8:  # Input VAT = Material Cost * 0.15
            cell.value = f"=G{row}*0.15"
            cell.number_format = sar_format
        elif col == 9:  # Transportation (user enters)
            cell.number_format = sar_format
        elif col == 10:  # Transport VAT = Transportation * 0.15
            cell.value = f"=I{row}*0.15"
            cell.number_format = sar_format
        elif col == 11:  # Total COGS = Material + Input VAT + Transportation + Transport VAT
            cell.value = f"=G{row}+H{row}+I{row}+J{row}"
            cell.number_format = sar_format

        # SUMMARY SECTION
        elif col == 15:  # Gross Profit = Sales - Material Cost - Transportation
            cell.value = f"=C{row}-G{row}-I{row}"
            cell.number_format = sar_format
        elif col == 16:  # VAT Difference = Output VAT - Input VAT - Transport VAT
            cell.value = f"=D{row}-H{row}-J{row}"
            cell.number_format = sar_format
        elif col == 17:  # Zakat @2.5% = Gross Profit * 0.025
            cell.value = f"=O{row}*0.025"
            cell.number_format = sar_format
        elif col == 18:  # Yousef @7.5% = Gross Profit * 0.075
            cell.value = f"=O{row}*0.075"
            cell.number_format = sar_format
        elif col == 19:  # Everyone @17% = Gross Profit * 0.17
            cell.value = f"=O{row}*0.17"
            cell.number_format = sar_format
        elif col == 20:  # Charity @5% = Gross Profit * 0.05
            cell.value = f"=O{row}*0.05"
            cell.number_format = sar_format

# ========================== SHEET 3: VAT SUMMARY ==========================
ws3 = wb.create_sheet("VAT Summary", 2)

# Set column widths
widths_3 = [20, 18, 18, 18, 3, 20]
for i, width in enumerate(widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = width

# Headers
headers_3 = [
    "Total VAT Difference (Project)",
    "Total Output VAT (Expense)",
    "Net VAT Payable",
    "ZYVC (Zakat+Yousef+Charity+VAT)",
]

for col, header in enumerate(headers_3, 1):
    cell = ws3.cell(row=1, column=col)
    cell.value = header
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws3.row_dimensions[1].height = 25

# Row 2: Single data row with totals
row = 2

# Column A: Total VAT Difference from Project Costing (Column P)
cell = ws3.cell(row=row, column=1)
cell.value = "=SUM('Project Costing'!P2:P51)"
cell.number_format = sar_format
cell.border = BORDER
cell.alignment = Alignment(horizontal="right", vertical="center")
cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
cell.font = Font(bold=True)

# Column B: Total Output VAT from Expense Sheet (Column G)
cell = ws3.cell(row=row, column=2)
cell.value = "=SUM(Expense!G2:G51)"
cell.number_format = sar_format
cell.border = BORDER
cell.alignment = Alignment(horizontal="right", vertical="center")
cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
cell.font = Font(bold=True)

# Column C: Net VAT Payable = A - B
cell = ws3.cell(row=row, column=3)
cell.value = "=A2-B2"
cell.number_format = sar_format
cell.border = BORDER
cell.alignment = Alignment(horizontal="right", vertical="center")
cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
cell.font = Font(bold=True)

# Column D: ZYVC = (Zakat + Yousef + Charity) + Net VAT
cell = ws3.cell(row=row, column=4)
cell.value = "=(SUM('Project Costing'!Q2:Q51)+SUM('Project Costing'!R2:R51)+SUM('Project Costing'!T2:T51))+C2"
cell.number_format = sar_format
cell.border = BORDER
cell.alignment = Alignment(horizontal="right", vertical="center")
cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
cell.font = Font(bold=True)

# ========================== SHEET 4: BALANCE SHEET ==========================
ws4 = wb.create_sheet("Balance Sheet", 3)

# Set column widths
widths_4 = [25, 20, 20, 20]
for i, width in enumerate(widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = width

# Title
title_cell = ws4.cell(row=1, column=1)
title_cell.value = "BALANCE SHEET - THE BLACK ARROW COMPANY"
title_cell.font = Font(bold=True, size=12, color=WHITE)
NAVY_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
title_cell.fill = NAVY_FILL
ws4.merge_cells('A1:D1')
ws4.row_dimensions[1].height = 25

# Section: INCOME
ws4.cell(row=3, column=1).value = "INCOME SECTION"
ws4.cell(row=3, column=1).font = Font(bold=True, size=11, color=WHITE)
ws4.cell(row=3, column=1).fill = SECTION_FILL
ws4.row_dimensions[3].height = 20

income_headers = ["Description", "Amount", "VAT", "Total"]
for col, header in enumerate(income_headers, 1):
    cell = ws4.cell(row=4, column=col)
    cell.value = header
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

income_items = [
    ("Total Sales Revenue", "=SUM('Project Costing'!C2:C51)", "=SUM('Project Costing'!D2:D51)", "=B5+C5"),
]

for idx, (item, amount_formula, vat_formula, total_formula) in enumerate(income_items, 5):
    ws4.cell(row=idx, column=1).value = item
    ws4.cell(row=idx, column=1).font = Font(bold=True)
    ws4.cell(row=idx, column=1).border = BORDER

    for col, formula in enumerate([amount_formula, vat_formula, total_formula], 2):
        cell = ws4.cell(row=idx, column=col)
        cell.value = formula
        cell.number_format = sar_format
        cell.border = BORDER

# Section: EXPENSES
exp_row = 8
ws4.cell(row=exp_row, column=1).value = "EXPENSE SECTION"
ws4.cell(row=exp_row, column=1).font = Font(bold=True, size=11, color=WHITE)
ws4.cell(row=exp_row, column=1).fill = SECTION_FILL
ws4.row_dimensions[exp_row].height = 20

expense_headers = ["Description", "Amount", "VAT", "Total"]
for col, header in enumerate(expense_headers, 1):
    cell = ws4.cell(row=exp_row+1, column=col)
    cell.value = header
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

expense_items = [
    ("Company Expenses", "=SUM(Expense!F2:F51)", "=SUM(Expense!G2:G51)", "=B10+C10"),
    ("Project COGS", "=SUM('Project Costing'!G2:G51)", "=SUM('Project Costing'!H2:H51)", "=B11+C11"),
]

for idx, (item, amount_formula, vat_formula, total_formula) in enumerate(expense_items, 10):
    ws4.cell(row=idx, column=1).value = item
    ws4.cell(row=idx, column=1).font = Font(bold=True)
    ws4.cell(row=idx, column=1).border = BORDER

    for col, formula in enumerate([amount_formula, vat_formula, total_formula], 2):
        cell = ws4.cell(row=idx, column=col)
        cell.value = formula
        cell.number_format = sar_format
        cell.border = BORDER

# Total Expenses
ws4.cell(row=12, column=1).value = "TOTAL EXPENSES"
ws4.cell(row=12, column=1).font = TOTAL_FONT
ws4.cell(row=12, column=1).fill = TOTAL_FILL
ws4.cell(row=12, column=1).border = BORDER

for col in range(2, 5):
    cell = ws4.cell(row=12, column=col)
    cell.value = f"=SUM(B10:B{11})" if col == 2 else (f"=SUM(C10:C{11})" if col == 3 else f"=SUM(D10:D{11})")
    cell.number_format = sar_format
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    cell.border = BORDER

# Section: PROFIT DISTRIBUTION
dist_row = 15
ws4.cell(row=dist_row, column=1).value = "PROFIT DISTRIBUTION"
ws4.cell(row=dist_row, column=1).font = Font(bold=True, size=11, color=WHITE)
ws4.cell(row=dist_row, column=1).fill = SECTION_FILL
ws4.row_dimensions[dist_row].height = 20

dist_headers = ["Partner/Purpose", "Amount", "Percentage", "Total"]
for col, header in enumerate(dist_headers, 1):
    cell = ws4.cell(row=dist_row+1, column=col)
    cell.value = header
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER

dist_items = [
    ("Zakat (2.5%)", "=SUM('Project Costing'!Q2:Q51)", "2.5%", "=B17"),
    ("Yousef Share (7.5%)", "=SUM('Project Costing'!R2:R51)", "7.5%", "=B18"),
    ("Everyone Share (17%)", "=SUM('Project Costing'!S2:S51)", "17%", "=B19"),
    ("Charity (5%)", "=SUM('Project Costing'!T2:T51)", "5%", "=B20"),
]

for idx, (partner, amount_formula, percentage, total_formula) in enumerate(dist_items, 17):
    ws4.cell(row=idx, column=1).value = partner
    ws4.cell(row=idx, column=1).border = BORDER

    cell = ws4.cell(row=idx, column=2)
    cell.value = amount_formula
    cell.number_format = sar_format
    cell.border = BORDER

    cell = ws4.cell(row=idx, column=3)
    cell.value = percentage
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER

    cell = ws4.cell(row=idx, column=4)
    cell.value = total_formula
    cell.number_format = sar_format
    cell.border = BORDER

# Total Distribution
ws4.cell(row=21, column=1).value = "TOTAL DISTRIBUTION"
ws4.cell(row=21, column=1).font = TOTAL_FONT
ws4.cell(row=21, column=1).fill = TOTAL_FILL
ws4.cell(row=21, column=1).border = BORDER

for col in [2, 4]:
    cell = ws4.cell(row=21, column=col)
    cell.value = f"=SUM(B17:B20)" if col == 2 else f"=SUM(D17:D20)"
    cell.number_format = sar_format
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    cell.border = BORDER

# Section: SUMMARY
summary_row = 24
ws4.cell(row=summary_row, column=1).value = "NET SUMMARY"
ws4.cell(row=summary_row, column=1).font = Font(bold=True, size=11, color=WHITE)
ws4.cell(row=summary_row, column=1).fill = SECTION_FILL
ws4.row_dimensions[summary_row].height = 20

summary_items = [
    ("Gross Profit (Total Sales - Total COGS)", "=B5-B11"),
    ("Net VAT Payable", "='VAT Summary'!G52"),
    ("Net Profit After Distribution", "=B25-B21-B26"),
]

for idx, (item, formula) in enumerate(summary_items, 25):
    ws4.cell(row=idx, column=1).value = item
    ws4.cell(row=idx, column=1).font = Font(bold=True)
    ws4.cell(row=idx, column=1).border = BORDER

    cell = ws4.cell(row=idx, column=2)
    cell.value = formula
    cell.number_format = sar_format
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    cell.border = BORDER

# Save file
file_path = r"C:\Users\Afzal\Downloads\Black_Arrow_Accounting_System.xlsx"
wb.save(file_path)
print(f"[OK] Professional Google Sheets-compatible workbook created: {file_path}")
print("\n[SHEETS CREATED - 4 Total]:")
print("  1. EXPENSE - Track all company expenses with VAT")
print("  2. PROJECT COSTING - Track sales, purchases, and profit distribution")
print("  3. VAT SUMMARY - Calculate net VAT payable to government")
print("  4. BALANCE SHEET - Complete financial summary")
print("\n[FEATURES]:")
print("  - All formulas are auto-linked between sheets!")
print("  - Simply enter: Date, Invoice Ref, Amounts -> Everything calculates automatically")
print("  - ZYVC = Zakat + Yousef + Charity + Final VAT Payable")
