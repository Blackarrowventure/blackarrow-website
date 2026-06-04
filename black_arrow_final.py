from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Black Arrow"

# Define colors and styles
PRIMARY_COLOR = "0F3460"
ACCENT_COLOR = "F59E0B"
HEADER_FILL = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUMMARY_FILL = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
SUMMARY_FONT = Font(bold=True, color="FFFFFF", size=10)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
col_widths = [5, 12, 15, 15, 15, 3, 15, 12, 15, 15, 12, 12, 15, 3, 15, 15, 15, 15, 15, 15]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Headers
headers = [
    "SN", "INV Ref", "Sales", "Output VAT", "Total Sales", " ",
    "Material cost", "Input VAT", "Transportation", "Transport VAT", "Total COGS", "P.O No",
    "Supplier ref", " ", "Gross Profit", "VAT Difference", "Zakat @2.5%", "Yousef Profit 7.5%",
    "Everyone Profit 17%", "Charity 5%"
]

# Add headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.row_dimensions[1].height = 30

# SAR currency format
sar_format = '"SAR"#,##0.00'

# Add 50 data rows with formulas
for row in range(2, 52):
    for col in range(1, 21):
        cell = ws.cell(row=row, column=col)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center")

        # Column assignments
        if col == 1:  # SN
            cell.value = row - 1
            cell.alignment = Alignment(horizontal="center", vertical="center")

        elif col == 3:  # Sales (user enters)
            cell.number_format = sar_format

        elif col == 4:  # Output VAT = Sales * 0.15
            cell.value = f"=C{row}*0.15"
            cell.number_format = sar_format

        elif col == 5:  # Total Sales = Sales + Output VAT
            cell.value = f"=C{row}+D{row}"
            cell.number_format = sar_format

        elif col == 7:  # Material cost (user enters)
            cell.number_format = sar_format

        elif col == 8:  # Input VAT = Material cost * 0.15
            cell.value = f"=G{row}*0.15"
            cell.number_format = sar_format

        elif col == 9:  # Transportation (user enters)
            cell.number_format = sar_format

        elif col == 10:  # Transport VAT = Transportation * 0.15
            cell.value = f"=I{row}*0.15"
            cell.number_format = sar_format

        elif col == 11:  # Total COGS = Material cost + Input VAT + Transportation + Transport VAT
            cell.value = f"=G{row}+H{row}+I{row}+J{row}"
            cell.number_format = sar_format

        elif col == 15:  # Gross Profit = Sales - Material cost - Transportation
            cell.value = f"=C{row}-G{row}-I{row}"
            cell.number_format = sar_format

        elif col == 16:  # VAT Difference = Output VAT - Input VAT - Transport VAT
            cell.value = f"=D{row}-H{row}-J{row}"
            cell.number_format = sar_format

        elif col == 17:  # Zakat @2.5% = Gross Profit * 0.025
            cell.value = f"=O{row}*0.025"
            cell.number_format = sar_format

        elif col == 18:  # Yousef Profit 7.5% = Gross Profit * 0.075
            cell.value = f"=O{row}*0.075"
            cell.number_format = sar_format

        elif col == 19:  # Everyone Profit 17% = Gross Profit * 0.17
            cell.value = f"=O{row}*0.17"
            cell.number_format = sar_format

        elif col == 20:  # Charity 5% = Gross Profit * 0.05
            cell.value = f"=O{row}*0.05"
            cell.number_format = sar_format

# Save file
file_path = r"C:\Users\Afzal\Downloads\Black_Arrow_Accounting.xlsx"
wb.save(file_path)
print(f"✓ Excel file created successfully: {file_path}")
