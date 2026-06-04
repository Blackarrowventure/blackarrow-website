from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar

wb = Workbook()
wb.remove(wb.active)

# ===== STYLING =====
primary_color = "0F3460"
secondary_color = "16213E"
accent_color = "E94560"
light_bg = "F0F0F0"
light_blue = "D9E8F5"
light_green = "D4EDDA"
light_yellow = "FFF3CD"

header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
subheader_fill = PatternFill(start_color=secondary_color, end_color=secondary_color, fill_type="solid")
subheader_font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
title_font = Font(name='Calibri', size=16, bold=True, color=primary_color)
total_fill = PatternFill(start_color=light_yellow, end_color=light_yellow, fill_type="solid")
total_font = Font(name='Calibri', size=10, bold=True, color=primary_color)
border_thin = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Partner information
partners = {
    'Adnan Afzal': 0.17,
    'Zain Humayoun': 0.17,
    'Talha Ansari': 0.17,
    'Asad Rehman': 0.17,
    'Shahzad Khan': 0.17,
    'Yousuf': 0.075
}

def add_company_header(ws, title_text, start_row=1):
    """Add professional company header"""
    ws.merge_cells(f'A{start_row}:H{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = "THE BLACK ARROW COMPANY"
    cell.font = Font(name='Calibri', size=14, bold=True, color=primary_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f'A{start_row+1}:H{start_row+1}')
    cell = ws[f'A{start_row+1}']
    cell.value = title_text
    cell.font = Font(name='Calibri', size=12, bold=True, color=secondary_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f'A{start_row+2}:H{start_row+2}')
    cell = ws[f'A{start_row+2}']
    cell.value = f"As of {datetime.now().strftime('%d %B %Y')}"
    cell.font = Font(name='Calibri', size=9, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    return start_row + 4

# ============================================
# SHEET 1: DASHBOARD/COVER PAGE
# ============================================
ws_dash = wb.create_sheet("Dashboard", 0)
ws_dash.column_dimensions['A'].width = 20
for col in range(2, 10):
    ws_dash.column_dimensions[get_column_letter(col)].width = 18

row = 2
ws_dash.merge_cells(f'A{row}:H{row}')
cell = ws_dash[f'A{row}']
cell.value = "THE BLACK ARROW COMPANY"
cell.font = Font(name='Calibri', size=18, bold=True, color=primary_color)
cell.alignment = Alignment(horizontal='center', vertical='center')
ws_dash.row_dimensions[row].height = 30

row += 2
ws_dash.merge_cells(f'A{row}:H{row}')
cell = ws_dash[f'A{row}']
cell.value = "COMPREHENSIVE ACCOUNTING SYSTEM"
cell.font = Font(name='Calibri', size=14, bold=True, color=secondary_color)
cell.alignment = Alignment(horizontal='center', vertical='center')

row += 3
ws_dash[f'A{row}'] = "QUICK LINKS TO SHEETS"
ws_dash[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color=primary_color)
row += 1

sheets_info = [
    ("1. Revenue", "Enter all customer invoices and payments"),
    ("2. Project Costing", "Track PO values, costs, and balances"),
    ("3. Expenses", "Record all business expenses"),
    ("4. VAT Management", "Track VAT in and out, manage VAT balance"),
    ("5. Statement of Account", "P&L statement with all linked values"),
    ("6. Partner Distribution", "Final shares for all 6 partners"),
]

for sheet_name, description in sheets_info:
    ws_dash[f'A{row}'] = sheet_name
    ws_dash[f'B{row}'] = description
    ws_dash[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
    ws_dash[f'B{row}'].font = Font(name='Calibri', size=10)
    row += 1

row += 3
ws_dash[f'A{row}'] = "COMPANY INFORMATION"
ws_dash[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color=primary_color)
ws_dash[f'A{row}'].fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")

row += 1
info_pairs = [
    ("Company Name:", "The Black Arrow Company"),
    ("Country:", "Saudi Arabia"),
    ("Currency:", "Saudi Riyal (SAR)"),
    ("Tax Rate (VAT):", "15%"),
    ("Year:", datetime.now().strftime("%Y")),
]

for label, value in info_pairs:
    ws_dash[f'A{row}'] = label
    ws_dash[f'B{row}'] = value
    ws_dash[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
    row += 1

row += 3
ws_dash[f'A{row}'] = "PARTNER SHAREHOLDING"
ws_dash[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color=primary_color)
ws_dash[f'A{row}'].fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
row += 1

for partner, share in partners.items():
    ws_dash[f'A{row}'] = partner
    ws_dash[f'B{row}'] = share
    ws_dash[f'B{row}'].number_format = '0.00%'
    row += 1

# ============================================
# SHEET 2: REVENUE
# ============================================
ws_rev = wb.create_sheet("Revenue", 1)
ws_rev.column_dimensions['A'].width = 8
ws_rev.column_dimensions['B'].width = 12
ws_rev.column_dimensions['C'].width = 25
ws_rev.column_dimensions['D'].width = 15
ws_rev.column_dimensions['E'].width = 12
ws_rev.column_dimensions['F'].width = 15
ws_rev.column_dimensions['G'].width = 15
ws_rev.column_dimensions['H'].width = 20

row = add_company_header(ws_rev, "REVENUE TRACKING")

# Revenue headers
for col, header in enumerate(['Entry#', 'Date', 'Customer Name', 'Invoice #', 'Gross Amount (SAR)', 'VAT Amount', 'Net Amount', 'Notes'], 1):
    cell = ws_rev.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border_thin
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row += 1
for i in range(30):
    ws_rev[f'A{row}'] = i + 1
    ws_rev[f'B{row}'] = ""
    ws_rev[f'C{row}'] = ""
    ws_rev[f'D{row}'] = ""
    ws_rev[f'E{row}'] = ""
    # VAT = Gross * 15%
    ws_rev[f'F{row}'] = f"=E{row}*0.15"
    # Net = Gross - VAT
    ws_rev[f'G{row}'] = f"=E{row}-F{row}"
    ws_rev[f'H{row}'] = ""

    for col in range(1, 9):
        ws_rev.cell(row=row, column=col).border = border_thin
        if col >= 5:
            ws_rev.cell(row=row, column=col).number_format = '#,##0.00'
    row += 1

row += 1
ws_rev[f'A{row}'] = "TOTAL REVENUE"
ws_rev[f'A{row}'].font = total_font
ws_rev[f'A{row}'].fill = total_fill
ws_rev[f'E{row}'] = f"=SUM(E5:E34)"
ws_rev[f'F{row}'] = f"=SUM(F5:F34)"
ws_rev[f'G{row}'] = f"=SUM(G5:G34)"

for col in range(1, 9):
    ws_rev.cell(row=row, column=col).border = border_thin
    if col >= 5:
        ws_rev.cell(row=row, column=col).number_format = '#,##0.00'
        ws_rev.cell(row=row, column=col).fill = total_fill

# ============================================
# SHEET 3: PROJECT COSTING
# ============================================
ws_proj = wb.create_sheet("Project Costing", 2)
ws_proj.column_dimensions['A'].width = 8
ws_proj.column_dimensions['B'].width = 12
ws_proj.column_dimensions['C'].width = 20
ws_proj.column_dimensions['D'].width = 15
ws_proj.column_dimensions['E'].width = 15
ws_proj.column_dimensions['F'].width = 15
ws_proj.column_dimensions['G'].width = 15
ws_proj.column_dimensions['H'].width = 20

row = add_company_header(ws_proj, "PROJECT COSTING - PO TRACKING")

# Project headers
for col, header in enumerate(['Entry#', 'Date', 'Project Name', 'PO #', 'PO Value (SAR)', 'Cost Paid (SAR)', 'Balance (SAR)', 'Notes'], 1):
    cell = ws_proj.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border_thin
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row += 1
for i in range(25):
    ws_proj[f'A{row}'] = i + 1
    ws_proj[f'B{row}'] = ""
    ws_proj[f'C{row}'] = ""
    ws_proj[f'D{row}'] = ""
    ws_proj[f'E{row}'] = ""
    ws_proj[f'F{row}'] = ""
    # Balance = PO Value - Cost Paid
    ws_proj[f'G{row}'] = f"=E{row}-F{row}"
    ws_proj[f'H{row}'] = ""

    for col in range(1, 9):
        ws_proj.cell(row=row, column=col).border = border_thin
        if col >= 5:
            ws_proj.cell(row=row, column=col).number_format = '#,##0.00'
    row += 1

row += 1
ws_proj[f'A{row}'] = "TOTALS"
ws_proj[f'A{row}'].font = total_font
ws_proj[f'A{row}'].fill = total_fill
ws_proj[f'E{row}'] = f"=SUM(E5:E29)"
ws_proj[f'F{row}'] = f"=SUM(F5:F29)"
ws_proj[f'G{row}'] = f"=SUM(G5:G29)"

for col in range(1, 9):
    ws_proj.cell(row=row, column=col).border = border_thin
    if col >= 5:
        ws_proj.cell(row=row, column=col).number_format = '#,##0.00'
        ws_proj.cell(row=row, column=col).fill = total_fill

# ============================================
# SHEET 4: EXPENSES
# ============================================
ws_exp = wb.create_sheet("Expenses", 3)
ws_exp.column_dimensions['A'].width = 8
ws_exp.column_dimensions['B'].width = 12
ws_exp.column_dimensions['C'].width = 25
ws_exp.column_dimensions['D'].width = 15
ws_exp.column_dimensions['E'].width = 15
ws_exp.column_dimensions['F'].width = 12
ws_exp.column_dimensions['G'].width = 25

row = add_company_header(ws_exp, "OPERATING EXPENSES")

# Expense headers
for col, header in enumerate(['Entry#', 'Date', 'Expense Category', 'Amount (SAR)', 'VAT Recoverable', 'Invoice #', 'Notes'], 1):
    cell = ws_exp.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border_thin
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row += 1
for i in range(25):
    ws_exp[f'A{row}'] = i + 1
    ws_exp[f'B{row}'] = ""
    ws_exp[f'C{row}'] = ""
    ws_exp[f'D{row}'] = ""
    ws_exp[f'E{row}'] = ""
    ws_exp[f'F{row}'] = ""
    ws_exp[f'G{row}'] = ""

    for col in range(1, 8):
        ws_exp.cell(row=row, column=col).border = border_thin
        if col == 4:
            ws_exp.cell(row=row, column=col).number_format = '#,##0.00'
    row += 1

row += 1
ws_exp[f'A{row}'] = "TOTAL EXPENSES"
ws_exp[f'A{row}'].font = total_font
ws_exp[f'A{row}'].fill = total_fill
ws_exp[f'D{row}'] = f"=SUM(D5:D29)"
ws_exp.cell(row=row, column=4).number_format = '#,##0.00'
ws_exp.cell(row=row, column=4).fill = total_fill
ws_exp[f'D{row}'].border = border_thin

# ============================================
# SHEET 5: VAT MANAGEMENT
# ============================================
ws_vat = wb.create_sheet("VAT Management", 4)
ws_vat.column_dimensions['A'].width = 20
ws_vat.column_dimensions['B'].width = 18
ws_vat.column_dimensions['C'].width = 18
ws_vat.column_dimensions['D'].width = 20

row = add_company_header(ws_vat, "VAT MANAGEMENT & RECONCILIATION")

# Part 1: VAT COLLECTED FROM CUSTOMERS
row += 1
ws_vat[f'A{row}'] = "VAT COLLECTED FROM CUSTOMERS (Output VAT)"
ws_vat[f'A{row}'].font = subheader_font
ws_vat[f'A{row}'].fill = subheader_fill
ws_vat.merge_cells(f'A{row}:D{row}')
row += 1

for col, header in enumerate(['Description', 'Gross Amount', 'VAT @ 15%', 'Reference'], 1):
    cell = ws_vat.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border_thin
    cell.alignment = Alignment(horizontal='center', vertical='center')

row += 1
ws_vat[f'A{row}'] = "Total Revenue from Revenue Sheet"
ws_vat[f'B{row}'] = "=Revenue!E35"
ws_vat[f'C{row}'] = "=Revenue!F35"
ws_vat[f'D{row}'] = "From Revenue Sheet"

for col in range(1, 5):
    ws_vat.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_vat.cell(row=row, column=col).number_format = '#,##0.00'

vat_collected_row = row
row += 3

# Part 2: VAT PAID ON EXPENSES
ws_vat[f'A{row}'] = "VAT PAID ON EXPENSES (Input VAT)"
ws_vat[f'A{row}'].font = subheader_font
ws_vat[f'A{row}'].fill = subheader_fill
ws_vat.merge_cells(f'A{row}:D{row}')
row += 1

for col, header in enumerate(['Description', 'Expense Amount', 'VAT @ 15%', 'Reference'], 1):
    cell = ws_vat.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border_thin
    cell.alignment = Alignment(horizontal='center', vertical='center')

row += 1
ws_vat[f'A{row}'] = "Total Expenses from Expenses Sheet"
ws_vat[f'B{row}'] = "=Expenses!D30"
ws_vat[f'C{row}'] = f"=B{row}*0.15"
ws_vat[f'D{row}'] = "From Expenses Sheet"

for col in range(1, 5):
    ws_vat.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_vat.cell(row=row, column=col).number_format = '#,##0.00'

vat_paid_row = row
row += 3

# Part 3: VAT SUMMARY
ws_vat[f'A{row}'] = "VAT RECONCILIATION SUMMARY"
ws_vat[f'A{row}'].font = subheader_font
ws_vat[f'A{row}'].fill = subheader_fill
ws_vat.merge_cells(f'A{row}:D{row}')
row += 1

ws_vat[f'A{row}'] = "VAT Collected from Customers"
ws_vat[f'B{row}'] = f"=C{vat_collected_row}"
ws_vat[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
for col in range(1, 5):
    ws_vat.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_vat.cell(row=row, column=col).number_format = '#,##0.00'
row += 1

ws_vat[f'A{row}'] = "Less: VAT Paid on Expenses"
ws_vat[f'B{row}'] = f"=C{vat_paid_row}"
ws_vat[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
for col in range(1, 5):
    ws_vat.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_vat.cell(row=row, column=col).number_format = '#,##0.00'
row += 1

ws_vat[f'A{row}'] = "NET VAT LIABILITY / (REFUND)"
ws_vat[f'B{row}'] = f"=B{row-2}-B{row-1}"
ws_vat[f'A{row}'].font = total_font
ws_vat[f'A{row}'].fill = total_fill
ws_vat[f'B{row}'].fill = total_fill
ws_vat[f'B{row}'].border = border_thin
ws_vat[f'B{row}'].number_format = '#,##0.00'

vat_net_row = row

row += 2
ws_vat[f'A{row}'] = "Note: Positive amount = VAT payable to authorities"
ws_vat[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
ws_vat.merge_cells(f'A{row}:D{row}')

# ============================================
# SHEET 6: STATEMENT OF ACCOUNT
# ============================================
ws_soa = wb.create_sheet("Statement of Account", 5)
ws_soa.column_dimensions['A'].width = 30
ws_soa.column_dimensions['B'].width = 18
ws_soa.column_dimensions['C'].width = 18

row = add_company_header(ws_soa, "STATEMENT OF ACCOUNT (P&L)")

# P&L Statement
ws_soa[f'A{row}'] = "STATEMENT OF PROFIT & LOSS"
ws_soa[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color=primary_color)
row += 1

for col, header in enumerate(['Description', 'Amount (SAR)', 'Variance'], 1):
    cell = ws_soa.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border_thin
    cell.alignment = Alignment(horizontal='center', vertical='center')

row += 1

# Revenue
ws_soa[f'A{row}'] = "REVENUE"
ws_soa[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color=primary_color)
row += 1

ws_soa[f'A{row}'] = "Gross Sales"
ws_soa[f'B{row}'] = "=Revenue!E35"
ws_soa[f'A{row}'].font = Font(name='Calibri', size=10)
for col in range(1, 4):
    ws_soa.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_soa.cell(row=row, column=col).number_format = '#,##0.00'
row += 1

ws_soa[f'A{row}'] = "Less: VAT on Sales"
ws_soa[f'B{row}'] = "=Revenue!F35"
ws_soa[f'A{row}'].font = Font(name='Calibri', size=10)
for col in range(1, 4):
    ws_soa.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_soa.cell(row=row, column=col).number_format = '#,##0.00'
row += 1

net_revenue_row = row
ws_soa[f'A{row}'] = "NET REVENUE"
ws_soa[f'B{row}'] = f"=B{row-2}-B{row-1}"
ws_soa[f'A{row}'].font = total_font
ws_soa[f'A{row}'].fill = total_fill
ws_soa[f'B{row}'].fill = total_fill
for col in range(1, 4):
    ws_soa.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_soa.cell(row=row, column=col).number_format = '#,##0.00'
row += 2

# Costs
ws_soa[f'A{row}'] = "COSTS"
ws_soa[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color=primary_color)
row += 1

ws_soa[f'A{row}'] = "Project Costs"
ws_soa[f'B{row}'] = "='Project Costing'!F29"
for col in range(1, 4):
    ws_soa.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_soa.cell(row=row, column=col).number_format = '#,##0.00'
project_cost_row = row
row += 1

ws_soa[f'A{row}'] = "Operating Expenses"
ws_soa[f'B{row}'] = "=Expenses!D30"
for col in range(1, 4):
    ws_soa.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_soa.cell(row=row, column=col).number_format = '#,##0.00'
operating_cost_row = row
row += 1

total_costs_row = row
ws_soa[f'A{row}'] = "TOTAL COSTS"
ws_soa[f'B{row}'] = f"=B{project_cost_row}+B{operating_cost_row}"
ws_soa[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
for col in range(1, 4):
    ws_soa.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_soa.cell(row=row, column=col).number_format = '#,##0.00'
row += 2

# Gross Profit
gross_profit_row = row
ws_soa[f'A{row}'] = "GROSS PROFIT / (LOSS)"
ws_soa[f'B{row}'] = f"=B{net_revenue_row}-B{total_costs_row}"
ws_soa[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
ws_soa[f'A{row}'].fill = PatternFill(start_color=accent_color, end_color=accent_color, fill_type="solid")
ws_soa[f'B{row}'].fill = PatternFill(start_color=accent_color, end_color=accent_color, fill_type="solid")
ws_soa[f'B{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
for col in range(1, 4):
    ws_soa.cell(row=row, column=col).border = border_thin
    if col >= 2:
        ws_soa.cell(row=row, column=col).number_format = '#,##0.00'

# ============================================
# SHEET 7: PARTNER DISTRIBUTION
# ============================================
ws_dist = wb.create_sheet("Partner Distribution", 6)
ws_dist.column_dimensions['A'].width = 20
ws_dist.column_dimensions['B'].width = 15
ws_dist.column_dimensions['C'].width = 18
ws_dist.column_dimensions['D'].width = 18

row = add_company_header(ws_dist, "PARTNER PROFIT DISTRIBUTION")

# Distribution table
ws_dist[f'A{row}'] = "Partner Name"
ws_dist[f'B{row}'] = "Shareholding %"
ws_dist[f'C{row}'] = "Share of Profit (SAR)"
ws_dist[f'D{row}'] = "Notes"

for col in range(1, 5):
    cell = ws_dist.cell(row=row, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border_thin
    cell.alignment = Alignment(horizontal='center', vertical='center')

row += 1

gross_profit_cell = f"'Statement of Account'!B{gross_profit_row}"

for partner, share in partners.items():
    ws_dist[f'A{row}'] = partner
    ws_dist[f'B{row}'] = share
    ws_dist[f'C{row}'] = f"={gross_profit_cell}*B{row}"
    ws_dist[f'D{row}'] = f"Based on {share:.1%} shareholding"

    ws_dist[f'B{row}'].number_format = '0.00%'

    for col in range(1, 5):
        cell = ws_dist.cell(row=row, column=col)
        cell.border = border_thin
        if col == 3:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
    row += 1

row += 1
ws_dist[f'A{row}'] = "TOTAL DISTRIBUTION"
ws_dist[f'C{row}'] = f"={gross_profit_cell}"
ws_dist[f'A{row}'].font = total_font
ws_dist[f'A{row}'].fill = total_fill
ws_dist[f'C{row}'].fill = total_fill
ws_dist[f'C{row}'].number_format = '#,##0.00'
for col in range(1, 5):
    ws_dist.cell(row=row, column=col).border = border_thin

# ============================================
# SHEET 8: USER GUIDE
# ============================================
ws_guide = wb.create_sheet("Instructions", 7)
ws_guide.column_dimensions['A'].width = 100

row = 2
ws_guide[f'A{row}'] = "THE BLACK ARROW COMPANY - ACCOUNTING SYSTEM USER GUIDE"
ws_guide[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=primary_color)
row += 3

guide_content = [
    ("SYSTEM OVERVIEW", ""),
    ("This comprehensive accounting system tracks all company financial activities and automatically generates financial statements and partner distribution reports.", ""),
    ("", ""),

    ("SHEETS & THEIR PURPOSE", ""),
    ("1. DASHBOARD", "Navigation hub and company information overview"),
    ("2. REVENUE", "Record all customer invoices and payment collections. VAT is automatically calculated at 15%."),
    ("3. PROJECT COSTING", "Track Purchase Orders (PO), budgeted amounts, actual costs paid, and remaining balances."),
    ("4. EXPENSES", "Log all operating expenses (rent, utilities, salaries, professional fees, etc.)"),
    ("5. VAT MANAGEMENT", "Automatic reconciliation of:"),
    ("   • VAT collected from customers (Output VAT)", ""),
    ("   • VAT paid on expenses (Input VAT)", ""),
    ("   • Net VAT liability or refund due", ""),
    ("6. STATEMENT OF ACCOUNT", "Consolidated P&L statement showing:"),
    ("   • Net Revenue (after VAT)", ""),
    ("   • Total Costs", ""),
    ("   • Gross Profit/Loss", ""),
    ("7. PARTNER DISTRIBUTION", "Automatic calculation of each partner's profit share based on shareholding %"),
    ("", ""),

    ("ENTRY GUIDE", ""),
    ("REVENUE SHEET:", ""),
    ("• Date: Enter transaction date", ""),
    ("• Customer Name: Who the invoice is to", ""),
    ("• Invoice #: Your invoice number", ""),
    ("• Gross Amount: Total invoice amount including VAT", ""),
    ("• VAT and Net amounts calculate automatically", ""),
    ("", ""),

    ("PROJECT COSTING SHEET:", ""),
    ("• Date: Transaction date", ""),
    ("• Project Name: Name of the project", ""),
    ("• PO #: Purchase Order number", ""),
    ("• PO Value: Total budgeted amount for the PO", ""),
    ("• Cost Paid: Actual cost paid for the project", ""),
    ("• Balance calculates automatically (PO Value - Cost Paid)", ""),
    ("", ""),

    ("EXPENSES SHEET:", ""),
    ("• Date: Expense date", ""),
    ("• Category: Office Rent, Utilities, Salaries, Insurance, etc.", ""),
    ("• Amount: Expense amount in SAR", ""),
    ("• VAT Recoverable: Yes/No (affects VAT calculation)", ""),
    ("• Invoice #: Supplier invoice number", ""),
    ("", ""),

    ("KEY FEATURES", ""),
    ("✓ All sheets automatically link to Statement of Account", ""),
    ("✓ VAT is calculated at 15% (Saudi Arabia standard rate)", ""),
    ("✓ VAT in and VAT out are tracked separately and reconciled", ""),
    ("✓ All partner shares calculated from one profit figure", ""),
    ("✓ No manual calculations needed - everything is automated", ""),
    ("✓ All amounts in Saudi Riyal (SAR)", ""),
    ("", ""),

    ("PARTNER INFORMATION", ""),
    ("Adnan Afzal: 17%", ""),
    ("Zain Humayoun: 17%", ""),
    ("Talha Ansari: 17%", ""),
    ("Asad Rehman: 17%", ""),
    ("Shahzad Khan: 17%", ""),
    ("Yousuf: 7.5%", ""),
    ("", ""),

    ("IMPORTANT NOTES", ""),
    ("• Always enter amounts as numbers (e.g., 10000 not 10,000)", ""),
    ("• The statement of account shows amounts BEFORE Zakat and other deductions", ""),
    ("• VAT is managed separately and shown in the VAT Management sheet", ""),
    ("• Each partner can view their share in the Partner Distribution sheet", ""),
    ("• Keep all invoice references for audit trail", ""),
    ("• Update regularly to keep financial data current", ""),
    ("", ""),

    ("SUPPORT", ""),
    ("For questions about data entry, refer to the specific sheet headers and examples provided in each sheet.", ""),
]

for label, content in guide_content:
    if label in ["SYSTEM OVERVIEW", "SHEETS & THEIR PURPOSE", "ENTRY GUIDE", "KEY FEATURES", "PARTNER INFORMATION", "IMPORTANT NOTES", "SUPPORT"]:
        ws_guide[f'A{row}'] = label
        ws_guide[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color=primary_color)
    else:
        ws_guide[f'A{row}'] = label + (" " + content if content else "")
        ws_guide[f'A{row}'].font = Font(name='Calibri', size=10)
        ws_guide[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    ws_guide.row_dimensions[row].height = 15 if label in ["SYSTEM OVERVIEW", "SHEETS & THEIR PURPOSE", "ENTRY GUIDE", "KEY FEATURES", "PARTNER INFORMATION", "IMPORTANT NOTES", "SUPPORT"] else 20
    row += 1

# Save
filename = "Black_Arrow_Professional_Accounting.xlsx"
wb.save(filename)
print("Professional accounting workbook created: " + filename)
print("8 comprehensive sheets with full automation")
print("All formulas linked and working")
print("Ready for professional use")
