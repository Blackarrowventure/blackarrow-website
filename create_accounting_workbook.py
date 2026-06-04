from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Define colors and styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(name='Calibri', size=10, bold=True, color="1F4E78")
total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
total_font = Font(name='Calibri', size=10, bold=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Partner information
partners = {
    'Adnan Afzal': 0.17,
    'Zain Humayoun': 0.17,
    'Talha Ansari': 0.17,
    'Asad Rehman': 0.17,
    'Shahzad Khan': 0.17,
    'Yousuf': 0.075
}

# ============================================
# SHEET 1: STATEMENT OF ACCOUNT (Main Sheet)
# ============================================
ws1 = wb.create_sheet("Statement of Account", 0)
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15

row = 1
ws1[f'A{row}'] = "THE BLACK ARROW COMPANY"
ws1[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
row += 1
ws1[f'A{row}'] = "STATEMENT OF ACCOUNT"
ws1[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
row += 1
ws1[f'A{row}'] = f"As at {datetime.now().strftime('%d %B %Y')}"
ws1[f'A{row}'].font = Font(name='Calibri', size=10, italic=True)
row += 2

# Account Summary
ws1[f'A{row}'] = "ACCOUNT SUMMARY"
ws1[f'A{row}'].font = subheader_font
ws1[f'A{row}'].fill = subheader_fill
row += 1

# Summary headers
for col, header in enumerate(['Description', 'Debit (SAR)', 'Credit (SAR)'], 1):
    cell = ws1.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')
row += 1

# Summary items
items = [
    'Total Revenue',
    'VAT - Claimable Expenses',
    'VAT - Non-Claimable Expenses',
    'Project Costing',
    'Official Non-Direct Costs',
    'Total Expenses',
    'NET PROFIT / (LOSS)'
]

summary_row_map = {}
for item in items:
    summary_row_map[item] = row
    ws1[f'A{row}'] = item
    ws1[f'A{row}'].font = Font(name='Calibri', size=10, bold=True if 'NET' in item else False)
    if 'NET' in item:
        ws1[f'A{row}'].fill = total_fill
    ws1[f'A{row}'].border = border

    # Add formulas
    if item == 'VAT - Claimable Expenses':
        ws1[f'B{row}'] = "=SUMIF('VAT General Expenses'!E:E,\"Claimable\",'VAT General Expenses'!D:D)"
    elif item == 'VAT - Non-Claimable Expenses':
        ws1[f'B{row}'] = "=SUMIF('VAT General Expenses'!E:E,\"Non-Claimable\",'VAT General Expenses'!D:D)"
    elif item == 'Project Costing':
        ws1[f'B{row}'] = "='Project Costing'!D22"
    elif item == 'Official Non-Direct Costs':
        ws1[f'B{row}'] = "='Official Costs Non-Direct'!D22"
    elif item == 'Total Expenses':
        exp_row = summary_row_map.get('VAT - Claimable Expenses', 5)
        ws1[f'B{row}'] = f"=B{exp_row}+B{exp_row+1}+B{exp_row+2}+B{exp_row+3}"
    elif item == 'NET PROFIT / (LOSS)':
        rev_row = summary_row_map.get('Total Revenue', 5)
        exp_row = summary_row_map.get('Total Expenses', 10)
        ws1[f'C{row}'] = f"=B{rev_row}-B{exp_row}"
        ws1[f'C{row}'].fill = total_fill

    if item != 'NET PROFIT / (LOSS)':
        if item != 'Total Revenue':
            ws1[f'B{row}'].number_format = '#,##0.00'
        ws1[f'B{row}'].border = border
        ws1[f'B{row}'].alignment = Alignment(horizontal='right')
    if item == 'NET PROFIT / (LOSS)':
        ws1[f'C{row}'].number_format = '#,##0.00'
        ws1[f'C{row}'].border = border
        ws1[f'C{row}'].alignment = Alignment(horizontal='right')
        ws1[f'C{row}'].font = Font(name='Calibri', size=10, bold=True)

    row += 1

row += 2

# ============================================
# PARTNER SUMMARY TABLE
# ============================================
ws1[f'A{row}'] = "PARTNER BALANCES SUMMARY"
ws1[f'A{row}'].font = subheader_font
ws1[f'A{row}'].fill = subheader_fill
row += 1

# Partner headers
ws1[f'A{row}'] = "Partner Name"
ws1[f'B{row}'] = "Shareholding %"
ws1[f'C{row}'] = "Share Amount (SAR)"
ws1[f'D{row}'] = "Less: VAT Portion"
ws1[f'E{row}'] = "Less: Zakat Portion"
ws1[f'F{row}'] = "Net Balance (SAR)"

for col in range(1, 7):
    cell = ws1.cell(row=row, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.column_dimensions[get_column_letter(col)].width = 18

partner_start_row = row + 1
row += 1
net_profit_ref = f"C{summary_row_map['NET PROFIT / (LOSS)']}"

for idx, (partner_name, shareholding) in enumerate(partners.items()):
    ws1[f'A{row}'] = partner_name
    ws1[f'B{row}'] = shareholding
    ws1[f'B{row}'].number_format = '0.00%'
    ws1[f'C{row}'] = f"=${net_profit_ref}*B{row}"
    ws1[f'D{row}'] = f"=C{row}*0.15"
    ws1[f'E{row}'] = f"=C{row}*0.025"
    ws1[f'F{row}'] = f"=C{row}-D{row}-E{row}"

    for col in range(1, 7):
        cell = ws1.cell(row=row, column=col)
        cell.border = border
        if col >= 3:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
    row += 1

# ============================================
# SHEET 2: VAT GENERAL EXPENSES
# ============================================
ws2 = wb.create_sheet("VAT General Expenses", 1)
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 25
ws2.column_dimensions['H'].width = 30

row = 1
ws2[f'A{row}'] = "VAT GENERAL EXPENSES"
ws2[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
ws2.merge_cells(f'A{row}:H{row}')
row += 2

# CLAIMABLE SECTION
ws2[f'A{row}'] = "CLAIMABLE VAT EXPENSES (Official Invoices Available)"
ws2[f'A{row}'].fill = subheader_fill
ws2[f'A{row}'].font = subheader_font
ws2.merge_cells(f'A{row}:H{row}')
row += 1

headers = ['Entry#', 'Date', 'Description', 'Amount (SAR)', 'VAT Type', 'Invoice #', 'Invoice Link', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row += 1
start_claimable = row

for i in range(15):
    ws2[f'A{row}'] = row - start_claimable + 1
    ws2[f'B{row}'] = ""
    ws2[f'C{row}'] = ""
    ws2[f'D{row}'] = ""
    ws2[f'E{row}'] = "Claimable"
    ws2[f'F{row}'] = ""
    ws2[f'G{row}'] = ""
    ws2[f'H{row}'] = ""

    ws2[f'D{row}'].number_format = '#,##0.00'

    for col in range(1, 9):
        ws2.cell(row=row, column=col).border = border
    row += 1

row += 1

# NON-CLAIMABLE SECTION
ws2[f'A{row}'] = "NON-CLAIMABLE VAT EXPENSES (No Official Invoices)"
ws2[f'A{row}'].fill = subheader_fill
ws2[f'A{row}'].font = subheader_font
ws2.merge_cells(f'A{row}:H{row}')
row += 1

for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row += 1
start_non_claimable = row

for i in range(15):
    ws2[f'A{row}'] = row - start_non_claimable + 1
    ws2[f'B{row}'] = ""
    ws2[f'C{row}'] = ""
    ws2[f'D{row}'] = ""
    ws2[f'E{row}'] = "Non-Claimable"
    ws2[f'F{row}'] = ""
    ws2[f'G{row}'] = ""
    ws2[f'H{row}'] = ""

    ws2[f'D{row}'].number_format = '#,##0.00'

    for col in range(1, 9):
        ws2.cell(row=row, column=col).border = border
    row += 1

row += 2

# Summary
ws2[f'A{row}'] = "TOTAL CLAIMABLE VAT"
ws2[f'A{row}'].font = total_font
ws2[f'A{row}'].fill = total_fill
ws2[f'D{row}'] = f"=SUMIF(E:E,\"Claimable\",D:D)"
ws2[f'D{row}'].number_format = '#,##0.00'
ws2[f'D{row}'].fill = total_fill
ws2[f'D{row}'].border = border

# ============================================
# SHEET 3: PROJECT COSTING
# ============================================
ws3 = wb.create_sheet("Project Costing", 2)
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 15
ws3.column_dimensions['F'].width = 12
ws3.column_dimensions['G'].width = 25

row = 1
ws3[f'A{row}'] = "PROJECT COSTING"
ws3[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
ws3.merge_cells(f'A{row}:G{row}')
row += 2

headers = ['Entry#', 'Date', 'Project Name', 'Cost Amount (SAR)', 'Cost Type', 'Invoice #', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row += 1

for i in range(20):
    ws3[f'A{row}'] = i + 1
    ws3[f'B{row}'] = ""
    ws3[f'C{row}'] = ""
    ws3[f'D{row}'] = ""
    ws3[f'E{row}'] = ""
    ws3[f'F{row}'] = ""
    ws3[f'G{row}'] = ""

    ws3[f'D{row}'].number_format = '#,##0.00'

    for col in range(1, 8):
        ws3.cell(row=row, column=col).border = border
    row += 1

row += 1

# Summary at row 22
ws3[f'A{row}'] = "TOTAL PROJECT COSTS"
ws3[f'A{row}'].font = total_font
ws3[f'A{row}'].fill = total_fill
ws3[f'D{row}'] = f"=SUM(D3:D22)"
ws3[f'D{row}'].number_format = '#,##0.00'
ws3[f'D{row}'].fill = total_fill
ws3[f'D{row}'].border = border

# ============================================
# SHEET 4: OFFICIAL COSTS NON-DIRECT
# ============================================
ws4 = wb.create_sheet("Official Costs Non-Direct", 3)
ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 15
ws4.column_dimensions['F'].width = 12
ws4.column_dimensions['G'].width = 25

row = 1
ws4[f'A{row}'] = "OFFICIAL NON-DIRECT COSTS"
ws4[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
ws4.merge_cells(f'A{row}:G{row}')
row += 2

ws4[f'A{row}'] = "Cost Categories: Office Rent, Utilities, Salaries, Insurance, Professional Fees, etc."
ws4[f'A{row}'].font = Font(name='Calibri', size=9, italic=True)
row += 1

headers = ['Entry#', 'Date', 'Cost Category', 'Amount (SAR)', 'Cost Type', 'Invoice #', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws4.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row += 1

for i in range(20):
    ws4[f'A{row}'] = i + 1
    ws4[f'B{row}'] = ""
    ws4[f'C{row}'] = ""
    ws4[f'D{row}'] = ""
    ws4[f'E{row}'] = ""
    ws4[f'F{row}'] = ""
    ws4[f'G{row}'] = ""

    ws4[f'D{row}'].number_format = '#,##0.00'

    for col in range(1, 8):
        ws4.cell(row=row, column=col).border = border
    row += 1

row += 1

# Summary at row 22
ws4[f'A{row}'] = "TOTAL NON-DIRECT COSTS"
ws4[f'A{row}'].font = total_font
ws4[f'A{row}'].fill = total_fill
ws4[f'D{row}'] = f"=SUM(D4:D23)"
ws4[f'D{row}'].number_format = '#,##0.00'
ws4[f'D{row}'].fill = total_fill
ws4[f'D{row}'].border = border

# ============================================
# SHEET 5: INSTRUCTIONS & NOTES
# ============================================
ws5 = wb.create_sheet("Instructions", 4)
ws5.column_dimensions['A'].width = 80

row = 1
ws5[f'A{row}'] = "ACCOUNTING WORKBOOK - USER GUIDE"
ws5[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
row += 2

instructions = [
    ("OVERVIEW:", "This workbook tracks all company expenses and generates a comprehensive Statement of Account."),
    ("", ""),
    ("SHEETS INCLUDED:", ""),
    ("1. Statement of Account", "Main summary sheet - shows total revenues, expenses, and partner balances (auto-calculated)"),
    ("2. VAT General Expenses", "Track VAT expenses separated into Claimable and Non-Claimable categories"),
    ("3. Project Costing", "Record direct project costs"),
    ("4. Official Costs Non-Direct", "Record indirect costs (office rent, utilities, salaries, insurance, etc.)"),
    ("", ""),
    ("HOW TO USE:", ""),
    ("Step 1:", "Enter your expense data in the relevant sheet (VAT, Project, or Non-Direct Costs)"),
    ("Step 2:", "Fill in: Date, Description, Amount (in SAR), and Invoice details"),
    ("Step 3:", "All calculations will automatically update in the Statement of Account"),
    ("Step 4:", "Partner balances are calculated automatically based on shareholding percentages"),
    ("", ""),
    ("IMPORTANT COLUMNS:", ""),
    ("Amount (SAR):", "Enter all amounts in Saudi Riyals. Use numbers only (no currency symbols)"),
    ("VAT Type:", "For VAT sheet - mark as 'Claimable' or 'Non-Claimable'"),
    ("Invoice Link:", "Paste URLs to cloud-stored invoices (Google Drive, OneDrive, etc.)"),
    ("Notes:", "Use for any additional information or clarifications"),
    ("", ""),
    ("PARTNER SHAREHOLDING:", ""),
    ("Adnan Afzal", "17%"),
    ("Zain Humayoun", "17%"),
    ("Talha Ansari", "17%"),
    ("Asad Rehman", "17%"),
    ("Shahzad Khan", "17%"),
    ("Yousuf", "7.5%"),
    ("", ""),
    ("CALCULATIONS:", ""),
    ("VAT Deduction:", "15% of net profit for each partner"),
    ("Zakat Deduction:", "2.5% of net profit for each partner"),
    ("Net Balance:", "Partner share - VAT - Zakat"),
    ("", ""),
    ("TIPS:", ""),
    ("- Keep invoice links organized for easy reference", ""),
    ("- Update regularly to avoid data entry errors", ""),
    ("- Use consistent date formats (DD/MM/YYYY)", ""),
    ("- Back up the file regularly", ""),
    ("- This workbook is Google Sheets compatible", ""),
]

for label, content in instructions:
    ws5[f'A{row}'] = label
    ws5[f'A{row}'].font = Font(name='Calibri', size=10, bold=True if label else False)
    if label and label.endswith(":") and not label.isupper():
        ws5[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color="1F4E78")
    row += 1

# Save workbook
filename = "Black_Arrow_Accounting_Workbook.xlsx"
wb.save(filename)
print(f"Excel workbook created: {filename}")
